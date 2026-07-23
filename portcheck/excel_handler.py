"""Excel (xlsx/xls) 导入/导出处理。"""

from __future__ import annotations

from pathlib import Path
from typing import Optional


def _is_header_row(values: list[str]) -> bool:
    """判断一行是否为标题行。"""
    if not values:
        return False
    first = str(values[0]).strip().lower()
    keywords = {"ip", "地址", "address", "host", "主机", "端口", "port", "描述", "description", "集合", "batch"}
    return first in keywords or any(kw in first for kw in keywords)


# ── 导入 ───────────────────────────────────────────────────


def parse_targets_excel(filepath: str | Path) -> tuple[list[dict], list[str]]:
    """从 Excel 文件导入目标。

    列顺序: IP地址, 端口, 描述, 集合名称 (第一行为标题则跳过)

    Returns:
        (targets, errors) — targets 为 [{"ip","port","description","batch_name"}, ...]
    """
    filepath = Path(filepath)
    ext = filepath.suffix.lower()
    targets: list[dict] = []
    errors: list[str] = []

    rows_data: list[list] = []

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(filepath))
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            rows_data.append([ws.cell_value(r, c) for c in range(ws.ncols)])
    else:
        from openpyxl import load_workbook
        wb = load_workbook(str(filepath), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))
        wb.close()

    for i, row in enumerate(rows_data, start=1):
        # 跳过完全空行
        values = [str(c).strip() if c is not None else "" for c in row]
        if all(v == "" for v in values):
            continue

        # 首行可能是标题
        if i == 1 and _is_header_row(values):
            continue

        if len(values) < 2:
            errors.append(f"第 {i} 行: 列数不足（需要 IP 和端口）")
            continue

        ip = values[0]
        port_str = values[1]
        desc = values[2] if len(values) > 2 else ""
        batch = values[3] if len(values) > 3 else ""

        try:
            port = int(float(port_str))  # Excel 可能把数字读成 float
        except (ValueError, TypeError):
            errors.append(f"第 {i} 行: 端口 '{port_str}' 无效")
            continue

        if not ip:
            errors.append(f"第 {i} 行: IP 地址为空")
            continue
        if not (1 <= port <= 65535):
            errors.append(f"第 {i} 行: 端口 {port} 超出范围")
            continue

        targets.append({"ip": ip, "port": port, "description": desc, "batch_name": batch})

    return targets, errors


# ── 导出 ───────────────────────────────────────────────────


def _format_row(values: list) -> list:
    """确保单元格值都是基础类型（非 numpy 等）。"""
    return [float(v) if isinstance(v, (int, float)) else str(v) for v in values]


def export_targets_to_excel(filepath: str | Path, targets: list[dict]) -> tuple[bool, str]:
    """导出目标列表到 Excel (.xlsx)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "目标列表"

        # 表头样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        headers = ["IP地址", "端口", "描述", "集合", "创建时间"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r, t in enumerate(targets, 2):
            ws.cell(row=r, column=1, value=t.get("ip", ""))
            ws.cell(row=r, column=2, value=int(t.get("port", 0)))
            ws.cell(row=r, column=3, value=t.get("description", ""))
            ws.cell(row=r, column=4, value=t.get("batch_name", ""))
            ws.cell(row=r, column=5, value=t.get("created_at", ""))

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 20

        wb.save(str(filepath))
        return True, ""
    except Exception as e:
        return False, str(e)


def export_results_to_excel(filepath: str | Path, results: list[dict]) -> tuple[bool, str]:
    """导出测试结果到 Excel (.xlsx)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "测试结果"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        headers = ["IP地址", "端口", "描述", "集合", "状态", "延迟(ms)", "错误信息", "检测时间"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r, t in enumerate(results, 2):
            success = t.get("success", False)
            row_fill = green_fill if success else red_fill

            for c, val in enumerate([
                t.get("ip", ""),
                int(t.get("port", 0)),
                t.get("description", ""),
                t.get("batch_name", ""),
                "连通" if success else "未连通",
                round(float(t.get("latency_ms", 0)), 1) if success else "",
                t.get("error_msg", ""),
                t.get("tested_at", ""),
            ], 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill = row_fill

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 20

        wb.save(str(filepath))
        return True, ""
    except Exception as e:
        return False, str(e)
